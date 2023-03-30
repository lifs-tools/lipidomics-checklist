import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import FormulaRule, DifferentialStyle, Rule
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.datavalidation import DataValidation
import json
import sqlite3



checked, unchecked = "[X]", "[  ]"

def create_condition_formula(condition, name_to_column, name_to_field, choice_to_field, row_num):
    global checked, unchecked
    
    conditions_split = condition.split("|")
    conditions_split = [condition_split.split("&") for condition_split in conditions_split]
    
    formatted_conditions = []
    for condition_or in conditions_split:
        formatted_conditions_or = []
        for condition in condition_or:
            split_val = "=" if condition.find("=") > -1 else "~"
            key, value = condition.split(split_val)
            field_type = None
            if "type" in name_to_field[key]:
                field_type = name_to_field[key]["type"]
                
            else:
                field_key = choice_to_field[key]
                field_type = name_to_field[field_key]["type"]
            
            if field_type == "text":
                value = value.strip('"')
                
            elif field_type == "number":
                value = str(value)
                
            elif field_type == "select":
                value = name_to_field[key]["label"]
                
            elif field_type == "multiple":
                value = checked
                
            key = "%s%i" % (name_to_column[key], row_num)
            formatted_conditions_or.append('%s%s"%s"' % (key, "<>" if split_val == "=" else "=", value))
        formatted_conditions.append("OR(%s)" % ", ".join(formatted_conditions_or) if len(formatted_conditions_or) > 1 else formatted_conditions_or[0])
    return "AND(%s)" % ", ".join(formatted_conditions) if len(formatted_conditions) > 1 else formatted_conditions[0]




def export_forms_to_worksheet(table_prefix, template, cursor, uid, main_entry_id):
    global checked, unchecked
    
    field_template = json.loads(open(template).read())
    workbook = Workbook()
    sheet = workbook.active


    name_to_column, name_to_field, name_to_data_validation, choice_to_field = {}, {}, {}, {}
    name_to_condition, name_to_select, multiple_names, columns = {}, {}, set(), []


    col_num = 1
    for page in field_template["pages"]:
        for field in page["content"]:
            if "name" not in field or "type" not in field: continue
            field_name = field["name"]
            cell_id = "%s1" % gcl(col_num)
        
            if field["type"] == "text":
                sheet.merge_cells("%s1:%s2" % (gcl(col_num), gcl(col_num)))
                cell = sheet[cell_id]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold = True)
                sheet[cell_id] = field["label"]
                sheet.column_dimensions[gcl(col_num)].width = len(field["label"]) + 4
                name_to_column[field_name] = gcl(col_num)
                name_to_field[field_name] = field
                columns.append(field_name)
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                col_num += 1
                
        
            elif field["type"] == "number":
                sheet.merge_cells("%s1:%s2" % (gcl(col_num), gcl(col_num)))
                cell = sheet[cell_id]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold = True)
                sheet[cell_id] = field["label"]
                sheet.column_dimensions[gcl(col_num)].width = len(field["label"]) + 4
                name_to_column[field_name] = gcl(col_num)
                name_to_field[field_name] = field
                columns.append(field_name)
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                col_num += 1
                
                
            elif field["type"] == "select":
                sheet.merge_cells("%s1:%s2" % (gcl(col_num), gcl(col_num)))
                cell = sheet[cell_id]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold = True)
                sheet[cell_id] = field["label"]
                sheet.column_dimensions[gcl(col_num)].width = len(field["label"]) + 4
                formula = ",".join(choice["label"] for choice in field["choice"])
                dv = DataValidation(type="list", formula1='"%s"' % formula, allow_blank = False)
                name_to_data_validation[field_name] = dv
                name_to_column[field_name] = gcl(col_num)
                name_to_select[field_name] = {choice["label"] for choice in field["choice"]}
                name_to_field[field_name] = field
                columns.append(field_name)
                for choice in field["choice"]:
                    name_to_column[choice["name"]] = gcl(col_num)
                    name_to_field[choice["name"]] = choice
                    choice_to_field[choice["name"]] = field_name
                col_num += 1
                
                
            elif field["type"] == "multiple":
                num_choice = len(field["choice"])
                sheet.merge_cells("%s1:%s1" % (gcl(col_num), gcl(col_num + num_choice - 1)))
                cell = sheet[cell_id]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold = True)
                sheet[cell_id] = field["label"]
                name_to_field[field_name] = field
                for choice in field["choice"]:
                    choice_name = choice["name"]
                    sheet.column_dimensions[gcl(col_num)].width = len(choice["label"]) + 4
                    cell_id = "%s2" % gcl(col_num)
                    sheet[cell_id] = choice["label"]
                    sheet[cell_id].font = Font(bold = True)
                    name_to_column[choice_name] = gcl(col_num)
                    dv = DataValidation(type="list", formula1='"%s,%s"' % (checked, unchecked), allow_blank = False)
                    name_to_data_validation[choice_name] = dv
                    name_to_field[choice_name] = choice
                    multiple_names.add(choice_name)
                    columns.append(choice_name)
                    choice_to_field[choice_name] = field_name
                    col_num += 1



        
    sql = "SELECT e.fields FROM %sentries AS e INNER JOIN %sconnect_sample AS s ON e.id = s.sample_form_entry_id WHERE s.main_form_entry_id = ? and e.user_id = ?;" % (table_prefix, table_prefix)
    cursor.execute(sql, (main_entry_id, uid))
    fields = [row["fields"] for row in cursor.fetchall()]


    row_num = 3
    for sample in fields:
        sample_name_to_field = {}
        
        for page in json.loads(sample)["pages"]:
            for field in page["content"]:
                if "name" not in field or "type" not in field: continue
                sample_name_to_field[field["name"]] = field
                if "choice" in field:
                    for choice in field["choice"]:
                        sample_name_to_field[choice["name"]] = choice
                        
                        
        for name in columns:
            col_name = name_to_column[name]
            cell_id = "%s%i" % (col_name, row_num)
            # add validations
            if name in name_to_data_validation: name_to_data_validation[name].add(cell_id)
            
            
            
            # add conditional formatting
            if name in name_to_condition:
                dxf = DifferentialStyle(fill = PatternFill(bgColor="DDDDDD"))
                r = Rule(type="expression", dxf = dxf)
                r.formula = [create_condition_formula(name_to_condition[name], name_to_column, name_to_field, choice_to_field, row_num)]
                sheet.conditional_formatting.add(cell_id, r)
            
            
            # fill column with values
            if name not in sample_name_to_field: continue
            field = sample_name_to_field[name]
            
            
            
            if name in multiple_names:
                sheet[cell_id] = checked if sample_name_to_field[name]["value"] == 1 else unchecked
                
            else:
                if field["type"] == "text" and name_to_field[name]["type"] == "text":
                    sheet[cell_id] = field["value"]
                
                elif field["type"] == "number" and name_to_field[name]["type"] == "number":
                    sheet[cell_id] = field["value"]
                
                elif field["type"] == "select" and name_to_field[name]["type"] == "select":
                    label = None
                    for choice in field["choice"]:
                        if choice["value"] == 1:
                            label = choice["label"]
                            break
                    sheet[cell_id] = label if name in name_to_select and label in name_to_select[name] else ""
            
        row_num += 1
                    
                    
    for _, dv in name_to_data_validation.items(): sheet.add_data_validation(dv)

    filename = "db/sample_list_%s.xlsx" % uid
    workbook.save(filename)

    return filename








def dict_factory(cursor, row):
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d

try:
    conn = sqlite3.connect("db/checklist.sqlite")
    conn.row_factory = dict_factory
    curr = conn.cursor()
except Exception as e:
    print("Error in dbconnect", e)
    exit()
    
export_forms_to_worksheet("TCrpQ_", "workflow-templates/sample.json", curr, 2, 156)
