import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import FormulaRule, DifferentialStyle, Rule
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from tempfile import NamedTemporaryFile
from json import loads as json_loads
from json import dumps as json_dumps
import base64
from datetime import datetime
from io import BytesIO
from FormsEnum import *
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

outgrey = "DDDDDD"
checked, unchecked = "[X]", "[  ]"
fake_comma = chr(11794)

def create_condition_formula(condition, name_to_column, name_to_field, choice_to_field, row_num):
    global checked, unchecked, outgrey
    
    conditions_split = condition.split("|")
    conditions_split = [condition_split.split("&") for condition_split in conditions_split]
    
    formatted_conditions = []
    
    for condition_or in conditions_split:
        formatted_conditions_or = []
        for condition in condition_or:
            
            split_val = "=" if condition.find("=") > -1 else "~"
            key, value = condition.split(split_val)
            field_type = None
            if key in name_to_field and "type" in name_to_field[key]:
                field_type = name_to_field[key]["type"]
                
            elif key in choice_to_field:
                field_key = choice_to_field[key]
                field_type = name_to_field[field_key]["type"]
                
            else:
                continue
            
            if field_type == "text":
                value = value.strip('"')
                
            elif field_type == "number":
                value = str(value)
                
            elif field_type == "select":
                value = name_to_field[key]["label"]
                
            elif field_type == "multiple":
                value = checked
                
            key = "%s%i" % (name_to_column[key], row_num)
            formatted_conditions_or.append('%s%s"%s"' % (key, "<>" if split_val == "=" else "=", value))
                
        
        if len(formatted_conditions_or) > 0: formatted_conditions.append("OR(%s)" % ", ".join(formatted_conditions_or) if len(formatted_conditions_or) > 1 else formatted_conditions_or[0])
                
    
    if len(formatted_conditions) == 0: return ""
    return "AND(%s)" % ", ".join(formatted_conditions) if len(formatted_conditions) > 1 else formatted_conditions[0]




def create_table_condition_formula(condition, name_to_column, name_to_field, choice_to_field, row_num, sheet_name, lookup_field):
    global checked, unchecked, outgrey
    
    conditions_split = condition.split("|")
    conditions_split = [condition_split.split("&") for condition_split in conditions_split]
    
    formatted_conditions = []
    for condition_or in conditions_split:
        formatted_conditions_or = []
        for condition in condition_or:
            split_val = "=" if condition.find("=") > -1 else "~"
            key, value = condition.split(split_val)
            field_type = None
            if key in name_to_field and "type" in name_to_field[key]:
                field_type = name_to_field[key]["type"]
                
            elif key in choice_to_field:
                field_key = choice_to_field[key]
                field_type = name_to_field[field_key]["type"]
            
            else:
                continue
            
            if field_type == "text":
                value = value.strip('"')
                
            elif field_type == "number":
                value = str(value)
                
            elif field_type == "select":
                value = name_to_field[key]["label"]
                
            elif field_type == "multiple":
                value = checked
                
            key = "VLOOKUP(A%i, '%s'!%s, %i, FALSE)" % (row_num, sheet_name, lookup_field, name_to_column[key])
                
            #key = "%i%i" % (name_to_column[key], row_num)
            formatted_conditions_or.append('%s%s"%s"' % (key, "<>" if split_val == "=" else "=", value))
            
        if len(formatted_conditions_or) > 0: formatted_conditions.append("OR(%s)" % ", ".join(formatted_conditions_or) if len(formatted_conditions_or) > 1 else formatted_conditions_or[0])
        
    if len(formatted_conditions) > 0:
        formula = "AND(%s)" % ", ".join(formatted_conditions) if len(formatted_conditions) > 1 else formatted_conditions[0]
    else:
        formula = ""
        
    formula = "IF(ISBLANK(A%i), TRUE, %s)" % (row_num, formula)
    return formula




class TableData:
    def __init__(self, _name, _title):
        self.title = _title
        self.name = _name
        self.column_names = ["Form ID"]
        self.data = []
        self.condition = None
        
    def add_row(self, f_id, row):
        tokens = row.split("|")
        l = len(self.column_names) - 1
        for i in range(0, len(tokens), l):
            self.data.append([f_id] + tokens[i : i + l])


    def get_content(self):
        content = {}
        for row in self.data:
            if row[0] not in content: content[row[0]] = []
            content[row[0]] += row[1:]
            
        for i in content:
            content[i] = "|".join(content[i])
        return content
        
        

def export_forms_to_worksheet(template, fields, sheet_name):
    global checked, unchecked, outgrey
    
    field_template = json_loads(open(template).read())
    workbook = Workbook()
    sheet = workbook[workbook.sheetnames[0]]
    sheet.title = sheet_name

    name_to_column, name_to_field, name_to_data_validation, choice_to_field = {}, {}, {}, {}
    name_to_column_number = {}
    name_to_condition, name_to_select, multiple_names, columns = {}, {}, set(), []
    table_data = {}
    table_data_list = []

    col_num = 1
    sheet["A1"].value = "form_id"
    cell_id = "%s2" % gcl(col_num)
    sheet.merge_cells("%s2:%s3" % (gcl(col_num), gcl(col_num)))
    cell = sheet[cell_id]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold = True)
    sheet[cell_id] = "ID"
    col_num = 2
    sheet.row_dimensions[1].hidden = True
    
    for page in field_template["pages"]:
        for field in page["content"]:
            if "name" not in field or "type" not in field: continue
            field_name = field["name"]
            cell_id = "%s2" % gcl(col_num)
        
            if field["type"] == "text":
                sheet["%s1" % gcl(col_num)].value = field_name
                sheet.merge_cells("%s2:%s3" % (gcl(col_num), gcl(col_num)))
                cell = sheet[cell_id]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold = True)
                cell.value = field["label"]
                sheet.column_dimensions[gcl(col_num)].width = len(field["label"]) + 4
                name_to_column[field_name] = gcl(col_num)
                name_to_column_number[field_name] = col_num
                name_to_field[field_name] = field
                columns.append(field_name)
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                col_num += 1
                
        
            elif field["type"] == "number":
                sheet["%s1" % gcl(col_num)].value = field_name
                sheet.merge_cells("%s2:%s3" % (gcl(col_num), gcl(col_num)))
                cell = sheet[cell_id]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold = True)
                cell.value = field["label"]
                sheet.column_dimensions[gcl(col_num)].width = len(field["label"]) + 4
                name_to_column[field_name] = gcl(col_num)
                name_to_column_number[field_name] = col_num
                name_to_field[field_name] = field
                columns.append(field_name)
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                col_num += 1
                
                
            elif field["type"] == "select":
                sheet["%s1" % gcl(col_num)].value = field_name
                sheet.merge_cells("%s2:%s3" % (gcl(col_num), gcl(col_num)))
                cell = sheet[cell_id]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold = True)
                cell.value = field["label"]
                sheet.column_dimensions[gcl(col_num)].width = len(field["label"]) + 4
                formula = ",".join(choice["label"].replace(",", fake_comma) for choice in field["choice"])
                dv = DataValidation(type="list", formula1='"%s"' % formula, allow_blank = False)
                name_to_data_validation[field_name] = dv
                name_to_column[field_name] = gcl(col_num)
                name_to_column_number[field_name] = col_num
                name_to_select[field_name] = {choice["label"] for choice in field["choice"]}
                name_to_field[field_name] = field
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                columns.append(field_name)
                for choice in field["choice"]:
                    name_to_column[choice["name"]] = gcl(col_num)
                    name_to_column_number[choice["name"]] = col_num
                    name_to_field[choice["name"]] = choice
                    choice_to_field[choice["name"]] = field_name
                col_num += 1
                
                
            elif field["type"] == "multiple":
                num_choice = len(field["choice"])
                sheet.merge_cells("%s2:%s2" % (gcl(col_num), gcl(col_num + num_choice - 1)))
                cell = sheet[cell_id]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold = True)
                cell.value = field["label"]
                name_to_field[field_name] = field
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                ml = len(field["label"]) + 4
                scl = sum(len(choice["label"]) + 4 for choice in field["choice"])
                if ml > scl:
                    scl = -1
                    ml = int(ml / len(field["choice"]))
                for choice in field["choice"]:
                    choice_name = choice["name"]
                    sheet["%s1" % gcl(col_num)].value = choice_name
                    sheet.column_dimensions[gcl(col_num)].width = len(choice["label"]) + 4 if scl > -1 else ml
                    cell_id = "%s3" % gcl(col_num)
                    sheet[cell_id] = choice["label"]
                    sheet[cell_id].font = Font(bold = True)
                    name_to_column[choice_name] = gcl(col_num)
                    name_to_column_number[choice_name] = col_num
                    dv = DataValidation(type="list", formula1='"%s,%s"' % (checked, unchecked), allow_blank = False)
                    name_to_data_validation[choice_name] = dv
                    name_to_field[choice_name] = choice
                    multiple_names.add(choice_name)
                    columns.append(choice_name)
                    choice_to_field[choice_name] = field_name
                    col_num += 1


            elif field["type"] == "table":
                columns.append(field_name)
                name_to_column[field_name] = gcl(col_num)
                name_to_column_number[field_name] = col_num
                sheet["%s1" % gcl(col_num)].value = field_name
                sheet.merge_cells("%s2:%s3" % (gcl(col_num), gcl(col_num)))
                cell = sheet[cell_id]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold = True)
                cell.value = field["label"] + "\n" + "Data in other sheet"
                sheet.column_dimensions[gcl(col_num)].width = len(field["label"]) + 4
                td = TableData(field["name"], field["label"])
                name_to_field[field_name] = field
                table_data_list.append(td)
                table_data[field["label"]] = td
                for col_name in field["columns"].split("|"): td.column_names.append(col_name)
                if "condition" in field: td.condition = field["condition"]
                col_num += 1
                
    row_num = 4
    if len(fields) == 0:
        sheet["A%i" % row_num].value = "=ROW() - 3"
        for name in columns:
            
            col_name = name_to_column[name]
            cell_id = "%s%i" % (col_name, row_num)
            # add validations
            if name in name_to_data_validation: name_to_data_validation[name].add(cell_id)
                
            # add conditional formatting
            if name in name_to_condition:
                dxf = DifferentialStyle(fill = PatternFill(bgColor=outgrey))
                rule = Rule(type="expression", dxf = dxf)
                rule.formula = [create_condition_formula(name_to_condition[name], name_to_column, name_to_field, choice_to_field, row_num)]
                sheet.conditional_formatting.add(cell_id, rule)
        
    else:
        for form in fields:
            
            form_name_to_field = {}
            
            for page in json_loads(form)["pages"]:
                for field in page["content"]:
                    if "name" not in field or "type" not in field: continue
                    form_name_to_field[field["name"]] = field
                    if "choice" in field:
                        for choice in field["choice"]:
                            form_name_to_field[choice["name"]] = choice
                                
                sheet["A%i" % row_num].value = "=ROW() - 3"  
                
            for name in columns:       
                col_name = name_to_column[name]
                cell_id = "%s%i" % (col_name, row_num)
                # add validations
                if name in name_to_data_validation: name_to_data_validation[name].add(cell_id)
                       
                    
                # add conditional formatting
                if name in name_to_condition:
                        dxf = DifferentialStyle(fill = PatternFill(bgColor=outgrey))
                        rule = Rule(type="expression", dxf = dxf)
                        rule.formula = [create_condition_formula(name_to_condition[name], name_to_column, name_to_field, choice_to_field, row_num)]
                        sheet.conditional_formatting.add(cell_id, rule)
                
                    
                
                # fill column with values
                if name not in form_name_to_field: continue
                field = form_name_to_field[name]
                
                if name in multiple_names:
                    sheet[cell_id].value = checked if form_name_to_field[name]["value"] == 1 else unchecked
                    
                else:
                    if field["type"] == "text" and name_to_field[name]["type"] == "text":
                        sheet[cell_id].value = field["value"]
                    
                    elif field["type"] == "number" and name_to_field[name]["type"] == "number":
                        sheet[cell_id].value = field["value"]
                    
                    elif field["type"] == "select" and name_to_field[name]["type"] == "select":
                        label = None
                        for choice in field["choice"]:
                            if choice["value"] == 1:
                                label = choice["label"]
                                break
                        sheet[cell_id].value = label if name in name_to_select and label in name_to_select[name] else ""
                        
                    elif field["type"] == "table":
                        sheet[cell_id].fill = PatternFill(start_color=outgrey, end_color=outgrey, fill_type='solid')
                
            # check for table data
            for td in table_data_list:
                if td.name not in form_name_to_field: continue
                field = form_name_to_field[td.name]
                if len(field["value"]) == 0: continue
                td.add_row(row_num - 3, field["value"])
            
            row_num += 1
                
            

    for _, dv in name_to_data_validation.items(): sheet.add_data_validation(dv)
    for selection in sheet.views.sheetView[0].selection:
        selection.activeCell = "A2"
        selection.sqref = "A2"

    
    
    lookup_field = "$A$4:$%s$10000" % gcl(len(columns))
    for td in table_data_list:
        workbook.create_sheet(td.title)
        sheet = workbook[td.title]
        sheet.row_dimensions[1].hidden = True
        sheet["A1"].value = td.name
        for selection in sheet.views.sheetView[0].selection:
            selection.activeCell = "A2"
            selection.sqref = "A2"
        dv = DataValidation(type="list", formula1="'%s'!$A$4:$A$10000" % (sheet_name), allow_blank = False)
        sheet.add_data_validation(dv)
        
        # add column names
        for c, col_name in enumerate(td.column_names):
            cell = sheet["%s2" % gcl(c + 1)]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold = True)
            cell.value = col_name
            sheet.column_dimensions[gcl(c + 1)].width = len(col_name) + 4
    
        row_num = 3
        if len(td.data) > 0:
            for row in td.data:
                col_num = 1
                for val in row:
                    cell_id = "%s%i" % (gcl(col_num), row_num)
                    cell = sheet[cell_id]
                    cell.value = val
                    
                    if col_num > 1:
                        dxf = DifferentialStyle(fill = PatternFill(bgColor=outgrey))
                        rule = Rule(type="expression", dxf = dxf)
                        rule.formula = [create_table_condition_formula(td.condition, name_to_column_number, name_to_field, choice_to_field, row_num, sheet_name, lookup_field)]
                        sheet.conditional_formatting.add(cell_id, rule)
                    
                    col_num += 1
                    
                dv.add("A%i" % row_num)
                row_num += 1
                
        else:
            dv.add("A%i" % row_num)
            
            for col_num in range(2, len(td.column_names) + 1):
                cell_id = "%s%i" % (gcl(col_num), row_num)
                dxf = DifferentialStyle(fill = PatternFill(bgColor=outgrey))
                rule = Rule(type="expression", dxf = dxf)
                rule.formula = [create_table_condition_formula(td.condition, name_to_column_number, name_to_field, choice_to_field, row_num, sheet_name, lookup_field)]
                sheet.conditional_formatting.add(cell_id, rule)
            
            

    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        output_stream = BytesIO(tmp.read())
        
    output_stream = base64.b64encode(output_stream.getvalue())
    output_stream = str(output_stream, "utf-8")

    return output_stream
    







def process_condition(conditions_text, field_types):
    conditions = conditions_text.split("|")
    conditions = [condition.split("&") for condition in conditions]
    
    for i, condition_and in enumerate(conditions):
        remove_and = []
        for j, condition in enumerate(condition_and):
            split_sign = "=" if condition.find("=") > -1 else "~"
            key, value = condition.split(split_sign)
            field_type = field_types[key]
            if field_type == "text":
                value = value.strip('"')
            elif field_type == "number":
                value = float(value)
            elif field_type in {"select", "multiple"}:
                value = int(value)
            
            conditions[i][j] = [key, split_sign, value]
            
    
    return conditions



def import_forms_from_worksheet(template, file_base_64):
    global checked, unchecked, outgrey

    t = base64.b64decode(file_base_64)
    workbook = load_workbook(filename=BytesIO(t))
    sheet = workbook[workbook.sheetnames[0]]

    
    name_to_data_validation, label_to_name = {}, {}
    name_to_condition, multiple_names = {}, set()
    field_types, required_names = {}, []
    field_visible, choice_to_field, names_list = {}, {}, []
    table_fields, tables, table_names = {}, {}, set()
    

    field_template = json_loads(open(template).read())
    for page in field_template["pages"]:
        for field in page["content"]:
            if "name" not in field or "type" not in field: continue
            field_name = field["name"]
            field_types[field_name] = field["type"]
            field_visible[field_name] = True
            names_list.append(field_name)
        
            if field["type"] == "text":
                label_to_name[field["label"]] = field_name
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                if "required" in field and field["required"] == 1: required_names.append(field_name)
                
        
            elif field["type"] == "number":
                label_to_name[field["label"]] = field_name
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                if "required" in field and field["required"] == 1: required_names.append(field_name)
                
                
            elif field["type"] == "select":
                label_to_name[field["label"]] = field_name
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                if "required" in field and field["required"] == 1: required_names.append(field_name)
                for choice in field["choice"]:
                    field_types[choice["name"]] = field["type"]
                    choice_to_field[choice["name"]] = field_name
                
                
            elif field["type"] == "multiple":
                if "required" in field and field["required"] == 1: required_names.append(field_name)
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                for choice in field["choice"]:
                    choice_name = choice["name"]
                    multiple_names.add(choice_name)
                    label_to_name["%s---%s" % (field["label"], choice["label"])] = choice_name
                    field_types[choice_name] = field["type"]
                    choice_to_field[choice_name] = field_name
                    
            elif field["type"] == "table":
                label_to_name[field["label"]] = field_name
                if "required" in field and field["required"] == 1: required_names.append(field_name)
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                table_fields[field["label"]] = field
                td = TableData(field_name, field["label"])
                td.column_names += field["columns"].split("|")
                tables[field_name] = td
                table_names.add(field_name)
                
    
    for name in name_to_condition:
        name_to_condition[name] = process_condition(name_to_condition[name], field_types)
    
    # load all table data
    for sht in workbook.sheetnames[1:]:
        name = workbook[sht]["A1"].value
        if name not in tables: continue
    
        sheet_table = workbook[sht]
        td = tables[name]
        
        col_num = 2
        col_to_index = {}
        column_names = []
        while sheet_table["%s2" % gcl(col_num)].value != None:
            col_to_index[sheet_table["%s2" % gcl(col_num)].value] = -1
            column_names.append(sheet_table["%s2" % gcl(col_num)].value)
            col_num += 1
        
        for i, col_name in enumerate(td.column_names):
            if col_name in col_to_index: col_to_index[col_name] = i
        
        
        for sheet_row in sheet_table.iter_rows(min_row = 2, values_only = True):
            form_id = sheet_row[0]
            try: form_id = int(form_id)
            except: continue
            if form_id == None:
                continue
            
            row = [form_id] + [""] * len(col_to_index)
            for col_num, col_name in enumerate(column_names):
                value = sheet_row[col_num + 1]
                if col_name in col_to_index:
                    row[col_to_index[col_name]] = value if value != None else ""
                    found_entry = False
    
            td.data.append(row)
    
    for name in tables:
        tables[name] = tables[name].get_content()
    
    
    
    # load main data
    # determine column_id / label pair
    column_labels = []
    skipped_empty_cols = 0
    col_num = 1
    previous_label = None
    imported_forms = []
    unset_columns = {v for k, v in label_to_name.items()}
    while col_num < 10000:
        if skipped_empty_cols >= 5: break
        
        if sheet["%s1" % gcl(col_num)].value == None and sheet["%s2" % gcl(col_num)].value == None:
            skipped_empty_cols += 1
            col_num += 1
            continue
        
        skipped_empty_cols = 0
        
        name = sheet["%s1" % gcl(col_num)].value
        if name in field_types:
            column_labels.append([col_num - 1, name])
            unset_columns.remove(name)
            
        col_num += 1
    
    
            
    # go through rows
    for row_num, sheet_row in enumerate(sheet.iter_rows(min_row = 4, values_only = True)):
        found_entry = False
    
        form_id = row_num + 1
        name_to_field = {}
        is_complete = True
        unset_names = list(unset_columns)
        
        # load fresh template
        field_template = json_loads(open(template).read())
        for page in field_template["pages"]:
            for field in page["content"]:
                if "name" not in field or "type" not in field: continue
                name_to_field[field["name"]] = field
                    
                if field["type"] in {"select", "multiple"}:
                    for choice in field["choice"]: name_to_field[choice["name"]] = choice
        
        # add all information from row into fields
        for col_id, name in column_labels:
            value = sheet_row[col_id]
            
            if name in table_names:
                if name not in name_to_field: continue
                field = name_to_field[name]
                
                if field["type"] == "table":
                    if field["name"] in tables and form_id in tables[field["name"]]:
                        field["value"] = tables[field["name"]][form_id]
                        found_entry = True
                continue
            
            if value == None:
                unset_names.append(name)
                continue
            
            
            if name in multiple_names:
                if name not in name_to_field: continue
                field = name_to_field[name]
                found_entry = True
                field["value"] = 1 if value == checked else 0
            
            else:
                if name not in name_to_field: continue
                field = name_to_field[name]
                found_entry = True
                
                if field["type"] == "text":
                    field["value"] = value
                    
                elif field["type"] == "number":
                    try:
                        field["value"] = float(value)
                    except:
                        unset_names.append(name)
                    
                elif field["type"] == "select":
                    found_choice = False
                    value = value.replace(fake_comma, ",")
                    for choice in field["choice"]:
                        if choice["label"] == value:
                            choice["value"] = 1
                            found_choice = True
                        else:
                            choice["value"] = 0
                            
                    if not found_choice and len(field["choice"]) > 0:
                        field["choice"][0]["value"] = 1
                        unset_names.append(name)
                    
            
        # check if form is complete, i.e., all required fields without conditions are filled
        # as well as all fields where the conditions are met
        for field_name in names_list:
            
            
            if field_name not in name_to_condition: continue
            field_visible[field_name] = False
                
                
            for condition_and in name_to_condition[field_name]:
                condition_met = True
                for key, operator, value in condition_and:
                    if key not in name_to_field: continue
                    
                    if key not in choice_to_field:
                        conditional_field = field_name
                    else:
                        conditional_field = choice_to_field[key]
                    
                    condition_met &= (conditional_field in field_visible and field_visible[conditional_field]) and ((operator == "=" and name_to_field[key]["value"] == value) or (operator == "~" and name_to_field[key]["value"] != value))
                
                field_visible[field_name] |= condition_met

            
        for name in required_names:
            if name not in field_visible or not field_visible[name]: continue
            if name in unset_names:
                is_complete = False
                break
        
            field = name_to_field[name]
            
            if field["type"] == "multiple":
                any_set = sum([choice["value"] for choice in field["choice"]])
                if any_set == 0:
                    is_complete = False
                    break
            
            """
            if field["type"] == "text":
                if field["value"] == "":
                    is_complete = False
                    raise Exception("e: %s" % name)
                    break
            
            elif field["type"] == "number":
                if field["value"] == "" or field["value"] == None:
                    field["value"] = 0
                    break
            
            elif field["type"] == "multiple":
                any_set = sum([choice["value"] for choice in field["choice"]])
                if any_set == 0:
                    is_complete = False
                    break
            
            elif field["type"] == "table":
                if field["value"] == "" or field["value"] == None:
                    is_complete = False
                    field["value"] = ""
                    break
            """
                
        imported_forms.append([field_template, is_complete])
            
    return imported_forms
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
def validate_form_complete(form, form_type):
        
    name_to_condition, field_types, required_names = {}, {}, []
    field_visible, choice_to_field, names_list = {}, {}, []
    unset_names, name_to_field = [], {}
    

    for page in form["pages"]:
        required_names_page = []
        required_names.append(required_names_page)
        for field in page["content"]:
            if "name" not in field or "type" not in field: continue
            field_name = field["name"]
            name_to_field[field_name] = field
            field_types[field_name] = field["type"]
            field_visible[field_name] = True
            names_list.append(field_name)
        
            if field["type"] == "text":
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                if "required" in field and field["required"] == 1: required_names_page.append(field_name)
                if not type(field["value"]) == str:
                    field["value"] = ""
                    unset_names.append(field_name)
                    continue
                
                if "validate" in field:
                    import re
                    pattern = re.compile(field["validate"])
                    if type(pattern.match(field["value"])) == None:
                        unset_names.append(field_name)
                        
                if field["value"] == "":
                    unset_names.append(field_name)
        
        
            elif field["type"] == "number":
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                if "required" in field and field["required"] == 1: required_names_page.append(field_name)
                if not (type(field["value"]) == int or type(field["value"]) == float):
                    field["value"] = 0
                    unset_names.append(field_name)
                    continue
                if ("min" in field and field["min"] > field["value"]) or ("max" in field and field["max"] < field["value"]):
                    unset_names.append(field_name)
                    
            
                
            elif field["type"] == "select":
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                if "required" in field and field["required"] == 1: required_names_page.append(field_name)
                set_choice = False
                for choice in field["choice"]:
                    field_types[choice["name"]] = field["type"]
                    choice_to_field[choice["name"]] = field_name
                    name_to_field[choice["name"]] = choice
                    
                    if choice["value"] not in {0, 1}:
                        choice["value"]
                        continue
                    
                    if choice["value"] == 1:
                        if not set_choice: set_choice = True
                        else:
                            choice["value"] = 0
                            
                if not set_choice:
                    unset_names.append(field_name)
                    if len(field["choice"]) > 0: field["choice"][0]["value"] = 1
                    
                
            elif field["type"] == "multiple":
                if "required" in field and field["required"] == 1: required_names_page.append(field_name)
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                set_choices = 0
                for choice in field["choice"]:
                    choice_name = choice["name"]
                    field_types[choice_name] = field["type"]
                    name_to_field[choice["name"]] = choice
                    choice_to_field[choice_name] = field_name
                    if choice["value"] not in {0, 1}:
                        choice["value"]
                        continue
                    set_choices += choice["value"] == 1
                    
                if set_choices == 0:
                    unset_names.append(field_name)
                    
            elif field["type"] == "table":
                if "required" in field and field["required"] == 1: required_names_page.append(field_name)
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                
    
    for name in name_to_condition:
        name_to_condition[name] = process_condition(name_to_condition[name], field_types)
        
    
    fine_pages = len(required_names)
                    
    # check if form is complete, i.e., all required fields without conditions are filled
    # as well as all fields where the conditions are met
    for field_name in names_list:
        if field_name not in name_to_condition: continue
        field_visible[field_name] = False
            
            
        for condition_and in name_to_condition[field_name]:
            condition_met = True
            for key, operator, value in condition_and:
                if key not in name_to_field:
                    continue
                
                if key not in choice_to_field:
                    conditional_field = field_name
                else:
                    conditional_field = choice_to_field[key]
                
                condition_met &= (conditional_field in field_visible and field_visible[conditional_field]) and ((operator == "=" and name_to_field[key]["value"] == value) or (operator == "~" and name_to_field[key]["value"] != value))
                
            
            field_visible[field_name] |= condition_met
        
    
    for p, required_names_page in enumerate(required_names):
        for name in required_names_page:
            if name not in field_visible or not field_visible[name]: continue
            if name in unset_names:
                fine_pages = min(fine_pages, p)
                break
    """
    if form_type == FormType.LIPID_CLASS:
        lipid_name = ""
        for ch in form["pages"][0]["content"][0]["choice"]:
            if ch["value"] == 1:
                lipid_name = ch["label"]
                break
        raise Exception("ErrorCodes error %s %s %s %s" % (lipid_name, fine_pages, len(required_names), fine_pages == len(required_names)))
    """
    return fine_pages == len(required_names), fine_pages
        
        
        
        
        
        
        
        
def import_form_from_file(form, form_type, version):
    reference_form, completed = 0, True
    
    
    if form_type == FormType.CHECKLIST:
        reference_form = json_loads(open("workflow-templates/checklist.json").read())
        
    elif form_type == FormType.SAMPLE:
        reference_form = json_loads(open("workflow-templates/sample.json").read())
        
    elif form_type == FormType.LIPID_CLASS:
        reference_form = json_loads(open("workflow-templates/lipid-class.json").read())
        
    reference_form["version"] = version
    reference_form["creation_date"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    if "pages" not in reference_form or "pages" not in form:
        return reference_form, False
    
    try:
        ref_pages = {p["title"]: p["content"] for p in reference_form["pages"]}
        form_pages = {p["title"]: p["content"] for p in form["pages"]}
    except:
        return reference_form, False
     
    completed &= len(ref_pages) == len(form_pages)
    for ref_page_title, ref_page in ref_pages.items():
        if ref_page_title not in form_pages:
            completed = False
            continue
        
        form_page = form_pages[ref_page_title]
        
        try:
            ref_contents = {c["name"]: c for c in ref_page}
            form_contents = {c["name"]: c for c in form_page}
        except:
            completed = False
            continue
        
        completed &= len(ref_contents) == len(form_contents)
        for ref_content_name, ref_content in ref_contents.items():
            if ref_content_name not in form_contents:
                completed = False
                continue
            
            form_content = form_contents[ref_content_name]
            if "value" in ref_content:
                if "value" in form_content:
                    ref_content["value"] = form_content["value"]
                else:
                    completed = False
                
            if "choice" in ref_content:
                if "choice" not in form_content:
                    completed = False
                    
                else:
                    try:
                        ref_choices = {c["name"]: c for c in ref_content["choice"]}
                        form_choices = {c["name"]: c for c in form_content["choice"]}
                    except:
                        completed = False
                        continue
                    
                    completed &= len(ref_choices) == len(form_choices)
                    for ref_choice_name, ref_choice in ref_choices.items():
                        if ref_choice_name not in form_choices:
                            completed = False
                            continue
                        else:
                            
                            form_choice = form_choices[ref_choice_name]
                            if "value" in ref_choice:
                                if "value" in form_choice:
                                    ref_choice["value"] = form_choice["value"]
                                else:
                                    completed = False
                 
                           
    
    compl, fine_pages = validate_form_complete(reference_form, form_type)
    fine_pages = min(fine_pages, form["max_page"]) if "max_page" in form and type(form["max_page"]) == int else fine_pages
    completed &= compl
    if "max_page" not in reference_form: completed = False
    else: reference_form["max_page"] = fine_pages
    
    if fine_pages < len(form["pages"]) - 1: completed = False
    
        
    return reference_form, completed


if __name__ == "__main__":
    import sqlite3
    
    def dict_factory(cursor, row):
        d = {}
        for idx, col in enumerate(cursor.description):
            d[col[0]] = row[idx]
        return {col[0]: row[idx] for idx, col in enumerate(cursor.description)}

    try:
        conn = sqlite3.connect("db/checklist.sqlite")
        conn.row_factory = dict_factory
        cursor = conn.cursor()
    except Exception as e:
        print("Error in dbconnect", e)
        exit()
        
        
    test_case = "im-l"
    
    
    if test_case == "ex-s":
        sql = "SELECT e.fields FROM TCrpQ_entries AS e INNER JOIN TCrpQ_connect_sample AS s ON e.id = s.sample_form_entry_id WHERE s.main_form_entry_id = ? and e.user_id = ?;"
        cursor.execute(sql, (156, 2))
        fields = [row["fields"] for row in cursor.fetchall()]
        sheet_b64 = export_forms_to_worksheet("workflow-templates/sample.json", fields, "Sample forms")
        with open("test-sample.xlsx", "wb") as out:
            out.write(base64.b64decode(sheet_b64))
            
            
    if test_case == "ex-l":
        sql = "SELECT e.fields FROM TCrpQ_entries AS e INNER JOIN TCrpQ_connect_lipid_class AS s ON e.id = s.class_form_entry_id WHERE s.main_form_entry_id = ? and e.user_id = ?;"
        cursor.execute(sql, (156, 2))
        fields = [row["fields"] for row in cursor.fetchall()]
        sheet_b64 = export_forms_to_worksheet("workflow-templates/lipid-class.json", fields, "Lipid class forms")
        with open("test-lipid-class.xlsx", "wb") as out:
            out.write(base64.b64decode(sheet_b64))
        
        
    elif test_case == "im-s":
        with open("Sample-list.xlsx", "rb") as infile:
            file_base_64 = base64.b64encode(infile.read())
        imp_forms = import_forms_from_worksheet("workflow-templates/sample.json", file_base_64)
        
        
    elif test_case == "im-l":
        with open("test-lipid-class.xlsx", "rb") as infile:
            file_base_64 = base64.b64encode(infile.read())
        imp_forms = import_forms_from_worksheet("workflow-templates/lipid-class.json", file_base_64)
        
        for form in imp_forms: print(form[1])
    
    
    else:
        raise Exception("no such test case")
