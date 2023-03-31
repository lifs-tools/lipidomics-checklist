import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import FormulaRule, DifferentialStyle, Rule
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.worksheet.datavalidation import DataValidation
from tempfile import NamedTemporaryFile
import json
import sqlite3
import base64
from io import BytesIO



checked, unchecked = "[X]", "[  ]"
fake_comma = chr(11794)

def create_condition_formula(condition, name_to_column, name_to_field, choice_to_field, row_num):
    global checked, unchecked
    
    conditions_split = condition.split("|")
    conditions_split = [condition_split.split("&") for condition_split in conditions_split]
    
    formatted_conditions = []
    for condition_or in conditions_split:
        formatted_conditions_or = []
        for condition in condition_or:
            split_val = "=" if condition.find("=") > -1 else "~"
            key, value = condition.split(split_val)
            field_type = None
            if "type" in name_to_field[key]:
                field_type = name_to_field[key]["type"]
                
            else:
                field_key = choice_to_field[key]
                field_type = name_to_field[field_key]["type"]
            
            if field_type == "text":
                value = value.strip('"')
                
            elif field_type == "number":
                value = str(value)
                
            elif field_type == "select":
                value = name_to_field[key]["label"]
                
            elif field_type == "multiple":
                value = checked
                
            key = "%s%i" % (name_to_column[key], row_num)
            formatted_conditions_or.append('%s%s"%s"' % (key, "<>" if split_val == "=" else "=", value))
        formatted_conditions.append("OR(%s)" % ", ".join(formatted_conditions_or) if len(formatted_conditions_or) > 1 else formatted_conditions_or[0])
    return "AND(%s)" % ", ".join(formatted_conditions) if len(formatted_conditions) > 1 else formatted_conditions[0]




def export_forms_to_worksheet(table_prefix, template, cursor, uid, main_entry_id):
    global checked, unchecked
    
    field_template = json.loads(open(template).read())
    workbook = Workbook()
    sheet = workbook.active


    name_to_column, name_to_field, name_to_data_validation, choice_to_field = {}, {}, {}, {}
    name_to_condition, name_to_select, multiple_names, columns = {}, {}, set(), []


    col_num = 1
    for page in field_template["pages"]:
        for field in page["content"]:
            if "name" not in field or "type" not in field: continue
            field_name = field["name"]
            cell_id = "%s1" % gcl(col_num)
        
            if field["type"] == "text":
                sheet.merge_cells("%s1:%s2" % (gcl(col_num), gcl(col_num)))
                cell = sheet[cell_id]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold = True)
                sheet[cell_id] = field["label"]
                sheet.column_dimensions[gcl(col_num)].width = len(field["label"]) + 4
                name_to_column[field_name] = gcl(col_num)
                name_to_field[field_name] = field
                columns.append(field_name)
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                col_num += 1
                
        
            elif field["type"] == "number":
                sheet.merge_cells("%s1:%s2" % (gcl(col_num), gcl(col_num)))
                cell = sheet[cell_id]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold = True)
                sheet[cell_id] = field["label"]
                sheet.column_dimensions[gcl(col_num)].width = len(field["label"]) + 4
                name_to_column[field_name] = gcl(col_num)
                name_to_field[field_name] = field
                columns.append(field_name)
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                col_num += 1
                
                
            elif field["type"] == "select":
                sheet.merge_cells("%s1:%s2" % (gcl(col_num), gcl(col_num)))
                cell = sheet[cell_id]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold = True)
                sheet[cell_id] = field["label"]
                sheet.column_dimensions[gcl(col_num)].width = len(field["label"]) + 4
                formula = ",".join(choice["label"].replace(",", fake_comma) for choice in field["choice"])
                dv = DataValidation(type="list", formula1='"%s"' % formula, allow_blank = False)
                name_to_data_validation[field_name] = dv
                name_to_column[field_name] = gcl(col_num)
                name_to_select[field_name] = {choice["label"] for choice in field["choice"]}
                name_to_field[field_name] = field
                columns.append(field_name)
                for choice in field["choice"]:
                    name_to_column[choice["name"]] = gcl(col_num)
                    name_to_field[choice["name"]] = choice
                    choice_to_field[choice["name"]] = field_name
                col_num += 1
                
                
            elif field["type"] == "multiple":
                num_choice = len(field["choice"])
                sheet.merge_cells("%s1:%s1" % (gcl(col_num), gcl(col_num + num_choice - 1)))
                cell = sheet[cell_id]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold = True)
                sheet[cell_id] = field["label"]
                name_to_field[field_name] = field
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                for choice in field["choice"]:
                    choice_name = choice["name"]
                    sheet.column_dimensions[gcl(col_num)].width = len(choice["label"]) + 4
                    cell_id = "%s2" % gcl(col_num)
                    sheet[cell_id] = choice["label"]
                    sheet[cell_id].font = Font(bold = True)
                    name_to_column[choice_name] = gcl(col_num)
                    dv = DataValidation(type="list", formula1='"%s,%s"' % (checked, unchecked), allow_blank = False)
                    name_to_data_validation[choice_name] = dv
                    name_to_field[choice_name] = choice
                    multiple_names.add(choice_name)
                    columns.append(choice_name)
                    choice_to_field[choice_name] = field_name
                    col_num += 1



        
    sql = "SELECT e.fields FROM %sentries AS e INNER JOIN %sconnect_sample AS s ON e.id = s.sample_form_entry_id WHERE s.main_form_entry_id = ? and e.user_id = ?;" % (table_prefix, table_prefix)
    cursor.execute(sql, (main_entry_id, uid))
    fields = [row["fields"] for row in cursor.fetchall()]


    
    row_num = 3  
    if len(fields) == 0:
        for name in columns:
            col_name = name_to_column[name]
            cell_id = "%s%i" % (col_name, row_num)
            # add validations
            if name in name_to_data_validation: name_to_data_validation[name].add(cell_id)
            
            
            
            # add conditional formatting
            if name in name_to_condition:
                dxf = DifferentialStyle(fill = PatternFill(bgColor="DDDDDD"))
                r = Rule(type="expression", dxf = dxf)
                r.formula = [create_condition_formula(name_to_condition[name], name_to_column, name_to_field, choice_to_field, row_num)]
                sheet.conditional_formatting.add(cell_id, r)
        
        
    else:
        for sample in fields:
            sample_name_to_field = {}
            
            for page in json.loads(sample)["pages"]:
                for field in page["content"]:
                    if "name" not in field or "type" not in field: continue
                    sample_name_to_field[field["name"]] = field
                    if "choice" in field:
                        for choice in field["choice"]:
                            sample_name_to_field[choice["name"]] = choice
                            
                            
            for name in columns:
                col_name = name_to_column[name]
                cell_id = "%s%i" % (col_name, row_num)
                # add validations
                if name in name_to_data_validation: name_to_data_validation[name].add(cell_id)
                
                
                
                # add conditional formatting
                if name in name_to_condition:
                    dxf = DifferentialStyle(fill = PatternFill(bgColor="DDDDDD"))
                    r = Rule(type="expression", dxf = dxf)
                    r.formula = [create_condition_formula(name_to_condition[name], name_to_column, name_to_field, choice_to_field, row_num)]
                    sheet.conditional_formatting.add(cell_id, r)
                
                
                # fill column with values
                if name not in sample_name_to_field: continue
                field = sample_name_to_field[name]
                
                if name in multiple_names:
                    sheet[cell_id] = checked if sample_name_to_field[name]["value"] == 1 else unchecked
                    
                else:
                    if field["type"] == "text" and name_to_field[name]["type"] == "text":
                        sheet[cell_id] = field["value"]
                    
                    elif field["type"] == "number" and name_to_field[name]["type"] == "number":
                        sheet[cell_id] = field["value"]
                    
                    elif field["type"] == "select" and name_to_field[name]["type"] == "select":
                        label = None
                        for choice in field["choice"]:
                            if choice["value"] == 1:
                                label = choice["label"]
                                break
                        sheet[cell_id] = label if name in name_to_select and label in name_to_select[name] else ""
                
            row_num += 1
                    
                    
    for _, dv in name_to_data_validation.items(): sheet.add_data_validation(dv)

    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        output_stream = BytesIO(tmp.read())
        
    output_stream = base64.b64encode(output_stream.getvalue())
    output_stream = str(output_stream, "utf-8")
    
    return output_stream







def process_condition(conditions_text, field_types):
    conditions = conditions_text.split("|")
    conditions = [condition.split("&") for condition in conditions]
    for i, condition_and in enumerate(conditions):
        for j, condition in enumerate(condition_and):
            split_sign = "=" if condition.find("=") > -1 else "~"
            key, value = condition.split(split_sign)
            field_type = field_types[key]
            if field_type == "text":
                value = value.strip('"')
            elif field_type == "number":
                value = float(value)
            elif field_type in {"select", "multiple"}:
                value = int(value)
            
            conditions[i][j] = [key, split_sign, value]
    
    return conditions



def import_forms_from_worksheet(table_prefix, template, file_base_64, form_type, cursor, uid, main_entry_id):
    global checked, unchecked

    t = base64.b64decode(file_base_64)
    wb = load_workbook(filename=BytesIO(t))
    sheet = wb.active

    
    name_to_data_validation, label_to_name = {}, {}
    name_to_condition, multiple_names = {}, set()
    field_types, required_names = {}, []
    field_visible, choice_to_field, names_list = {}, {}, []

    field_template = json.loads(open(template).read())
    for page in field_template["pages"]:
        for field in page["content"]:
            if "name" not in field or "type" not in field: continue
            field_name = field["name"]
            field_types[field_name] = field["type"]
            field_visible[field_name] = True
            names_list.append(field_name)
        
            if field["type"] == "text":
                label_to_name[field["label"]] = field_name
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                if "required" in field and field["required"] == 1: required_names.append(field_name)
                
        
            elif field["type"] == "number":
                label_to_name[field["label"]] = field_name
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                if "required" in field and field["required"] == 1: required_names.append(field_name)
                
                
            elif field["type"] == "select":
                label_to_name[field["label"]] = field_name
                if "required" in field and field["required"] == 1: required_names.append(field_name)
                for choice in field["choice"]:
                    field_types[choice["name"]] = field["type"]
                    choice_to_field[choice["name"]] = field_name
                
                
            elif field["type"] == "multiple":
                if "required" in field and field["required"] == 1: required_names.append(field_name)
                if "condition" in field: name_to_condition[field_name] = field["condition"]
                for choice in field["choice"]:
                    choice_name = choice["name"]
                    multiple_names.add(choice_name)
                    label_to_name["%s---%s" % (field["label"], choice["label"])] = choice_name
                    field_types[choice_name] = field["type"]
                    choice_to_field[choice_name] = field_name
    
    
    for name in name_to_condition:
        name_to_condition[name] = process_condition(name_to_condition[name], field_types)
    
    
    
    
    # determine column_id / label pair
    column_labels = []
    skipped_empty_cols = 0
    col_num = 1
    previous_label = None
    imported_forms = []
    while col_num < 10000:
        if skipped_empty_cols >= 5: break
        
        if sheet["%s1" % gcl(col_num)].value == None and sheet["%s2" % gcl(col_num)].value == None:
            skipped_empty_cols += 1
            col_num += 1
            continue
        
        skipped_empty_cols = 0
        
        if sheet["%s2" % gcl(col_num)].value != None:
            if sheet["%s1" % gcl(col_num)].value != None:
                label = sheet["%s1" % gcl(col_num)].value + "---" + sheet["%s2" % gcl(col_num)].value
                previous_label = sheet["%s1" % gcl(col_num)].value
            else:
                label = previous_label + "---" + sheet["%s2" % gcl(col_num)].value
        
        else:
            label = sheet["%s1" % gcl(col_num)].value
            
        if label in label_to_name: column_labels.append([gcl(col_num), label])
        col_num += 1
        
        
    # go through rows
    row_num = 3
    skipped_empty_rows = 0
    while row_num < 10000:
        if skipped_empty_rows >= 5: break
        found_entry = False
        
        name_to_field = {}
        is_complete = True
        unset_names = []
        
        # load fresh template
        field_template = json.loads(open(template).read())
        for page in field_template["pages"]:
            for field in page["content"]:
                if "name" not in field or "type" not in field: continue
                name_to_field[field["name"]] = field
                    
                if field["type"] in {"select", "multiple"}:
                    for choice in field["choice"]: name_to_field[choice["name"]] = choice
        
        # add all information from row into fields
        for col_id, label in column_labels:
            name = label_to_name[label]
            value = sheet["%s%i" % (col_id, row_num)].value
            if value == None:
                unset_names.append(name)
                continue
            
            if name in multiple_names:
                if name not in name_to_field: continue
                field = name_to_field[name]
                found_entry = True
                field["value"] = 1 if value == checked else 0
            
            else:
                if name not in name_to_field: continue
                field = name_to_field[name]
                found_entry = True
                
                if field["type"] == "text":
                    field["value"] = value
                    
                elif field["type"] == "number":
                    try:
                        field["value"] = float(value)
                    except:
                        unset_names.append(name)
                    
                elif field["type"] == "select":
                    found_choice = False
                    value = value.replace(fake_comma, ",")
                    for choice in field["choice"]:
                        if choice["label"] == value:
                            choice["value"] = 1
                            found_choice = True
                        else:
                            choice["value"] = 0
                            
                    if not found_choice and len(field["choice"]) > 0:
                        field["choice"][0]["value"] = 1
                        unset_names.append(name)
                    
    
    
        if found_entry:
            
            # check if form is complete, i.e., all required fields without conditions are filled
            # as well as all fields where the conditions are met
            for field_name in names_list:
                if field_name not in name_to_condition: continue
                field_visible[field_name] = False
                for condition_and in name_to_condition[field_name]:
                    condition_met = True
                    for key, operator, value in condition_and:
                        conditional_field = choice_to_field[key]
                        condition_met &= (conditional_field in field_visible and field_visible[conditional_field]) and ((operator == "=" and name_to_field[key]["value"] == value) or (operator == "~" and name_to_field[key]["value"] != value))
                    
                    field_visible[field_name] |= condition_met
                    
            for name in unset_names:
                if name in field_visible and field_visible[name]: is_complete = False
                    
            
            for name in required_names:
                if name not in field_visible or not field_visible[name]: continue
            
                field = name_to_field[name]
                
                if field["type"] == "text":
                    if field["value"] == "":
                        is_complete = False
                        break
                
                elif field["type"] == "number":
                    if field["value"] == "" or field["value"] == None:
                        field["value"] = 0
                        print(row_num, field["name"])
                        break
                
                elif field["type"] == "multiple":
                    any_set = sum([choice["value"] for choice in field["choice"]])
                    if any_set == 0:
                        is_complete = False
                        break
            
            skipped_empty_rows = 0
            imported_forms.append([field_template, is_complete])
            
        else: skipped_empty_rows += 1
        row_num += 1
        
    return imported_forms
        
        
        
        
        
        


if __name__ == "__main__":
    def dict_factory(cursor, row):
        d = {}
        for idx, col in enumerate(cursor.description):
            d[col[0]] = row[idx]
        return d

    try:
        conn = sqlite3.connect("db/checklist.sqlite")
        conn.row_factory = dict_factory
        curr = conn.cursor()
    except Exception as e:
        print("Error in dbconnect", e)
        exit()
        
        
        
        
    with open("Sample-list.xlsx", "rb") as infile:
        file_base_64 = base64.b64encode(infile.read())
    imp_forms = import_forms_from_worksheet("TCrpQ_", "workflow-templates/sample.json", file_base_64, "sample", curr, 2, 156)
    
    for form in imp_forms: print(form[1])

